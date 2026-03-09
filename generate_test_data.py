"""
生成导入测试数据
使用 openpyxl 确保卡号以文本格式保存，避免科学计数法问题
"""
from openpyxl import Workbook

def generate_simple_test_data():
    """生成简单测试数据（3条记录）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 添加表头
    headers = ["卡号", "卡类型", "卡主姓名", "用户ID", "批次"]
    ws.append(headers)

    # 创建简单测试数据（使用 label 格式）
    test_data = [
        ["6217001234567890123", "疑似员工卡", "张三", "external_user_001", 1],
        ["6228481234567890456", "员工卡", "李四", "external_user_002", 1],
        ["1234567890123456", "赌客/会员", "王五", "wx_user_12345", 2],
    ]

    # 添加数据并设置卡号列为文本格式
    for row_data in test_data:
        ws.append(row_data)

    # 设置卡号列（A列）为文本格式
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        cell.number_format = '@'  # @ 表示文本格式

    # 保存文件
    output_path = "/Users/yipf/Desktop/测试导入任务功能.xlsx"
    wb.save(output_path)

    print(f"✅ 已生成简单测试文档: {output_path}")
    print(f"📊 总记录数: {len(test_data)}")
    return output_path

def generate_complex_test_data():
    """生成复杂测试数据（18条记录）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 添加表头
    headers = ["卡号", "卡类型", "卡主姓名", "用户ID", "批次"]
    ws.append(headers)

    # 创建测试数据（使用 label 格式）
    test_data = [
        # 正常数据 - 能匹配银行
        ["6217001234567890123", "疑似员工卡", "张三", "external_user_001", 1],
        ["6228481234567890456", "赌客/会员", "李四", "external_user_002", 1],
        ["6222021234567890789", "员工卡", "王五", "external_user_003", 1],

        # 无法匹配银行的卡号
        ["1234567890123456", "赌客/会员", "赵六", "wx_user_12345", 2],
        ["9999999999999999", "疑似员工卡", "钱七", "wx_user_67890", 2],

        # 重复卡号（第二次导入时会失败）
        ["6217001234567890123", "员工卡", "孙八", "external_user_004", 2],

        # 无效卡类型
        ["6225881234567890111", "invalid_type", "周九", "external_user_005", 2],

        # 缺少必填字段 - 卡号为空
        ["", "赌客/会员", "吴十", "external_user_006", 3],

        # 缺少必填字段 - 卡类型为空
        ["6217851234567890222", "", "郑十一", "external_user_007", 3],

        # 正常数据 - 不同批次
        ["6217001234567890333", "疑似员工卡", "冯十二", "external_user_008", 3],
        ["6228481234567890444", "赌客/会员", "陈十三", "external_user_009", 3],

        # 边界情况 - 最短卡号（16位）
        ["6217001234567890", "员工卡", "褚十四", "external_user_010", 4],

        # 边界情况 - 最长卡号（19位）
        ["6217001234567890555", "疑似员工卡", "卫十五", "external_user_011", 4],

        # 特殊字符测试
        ["6228481234567890666", "赌客/会员", "蒋十六（测试）", "external_user_012", 4],

        # 长文本测试
        ["6222021234567890777", "员工卡", "沈十七", "external_user_with_very_long_id_for_testing_purposes_123456789", 4],

        # 更多正常数据用于测试分页
        ["6217001234567890888", "疑似员工卡", "韩十八", "external_user_013", 5],
        ["6228481234567890999", "赌客/会员", "杨十九", "external_user_014", 5],
        ["6222021234567891000", "员工卡", "朱二十", "external_user_015", 5],
    ]

    # 添加数据并设置卡号列为文本格式
    for row_data in test_data:
        ws.append(row_data)

    # 设置卡号列（A列）为文本格式
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        cell.number_format = '@'  # @ 表示文本格式

    # 保存文件
    output_path = "/Users/yipf/Desktop/复杂导入测试数据.xlsx"
    wb.save(output_path)

    print(f"✅ 已生成复杂测试文档: {output_path}")
    print(f"📊 总记录数: {len(test_data)}")
    print(f"\n测试场景统计:")
    print(f"  - 正常可匹配银行: 9条")
    print(f"  - 无法匹配银行: 2条")
    print(f"  - 重复卡号: 1条")
    print(f"  - 无效卡类型: 1条")
    print(f"  - 缺少必填字段: 2条")
    print(f"  - 边界情况测试: 2条")
    print(f"  - 特殊字符测试: 1条")
    print(f"  - 长文本测试: 1条")

    return output_path

if __name__ == "__main__":
    print("="*60)
    print("  生成导入测试数据")
    print("="*60)
    print()

    # 生成简单测试数据
    print("1. 生成简单测试数据...")
    simple_path = generate_simple_test_data()
    print()

    # 生成复杂测试数据
    print("2. 生成复杂测试数据...")
    complex_path = generate_complex_test_data()
    print()

    print("="*60)
    print("  生成完成！")
    print("="*60)
    print()
    print("✨ 重要提示:")
    print("   - 卡号已设置为文本格式，不会出现科学计数法问题")
    print("   - 使用 openpyxl 库生成，确保格式正确")
    print()
    print("📁 生成的文件:")
    print(f"   - 简单测试: {simple_path}")
    print(f"   - 复杂测试: {complex_path}")
